'''单元格样式默认值'''
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

font = Font(name='Calibri',                 #字体名
                 size=11,                   #大小
                 bold=False,                #加粗
                 italic=False,              #倾斜
                 vertAlign=None,            #垂直对齐
                 underline='none',          #下划线
                 strike=False,              #删除线
                 color='FF000000')          #颜色
fill = PatternFill(fill_type=None,                  #填充背景
                 start_color='FFFFFFFF',            #开始颜色
                 end_color='FF000000')              #结束颜色
border = Border(left=Side(border_style=None,            #边框设置：单黑线：Side(border_style="thin", color="000000")
                           color='FF000000'),           #双黑线：Side(border_style="double", color="000000")
                 right=Side(border_style=None,
                            color='FF000000'),
                 top=Side(border_style="double",
                          color='FF000000'),
                 bottom=Side(border_style="thin",
                             color='000000'),
                 diagonal=Side(border_style=None,       #对角线
                               color='FF000000'),
                 diagonal_direction=0,                  #对角线方向
                 outline=Side(border_style="double",        #外框
                              color='000000'),
                 vertical=Side(border_style="thin",       #内部垂直框？
                               color='000000'),
                 horizontal=Side(border_style="thin",     #内部水平框？
                                color='000000')
                )
alignment=Alignment(horizontal='general',       #水平对齐方式
                     vertical='bottom',         #垂直对齐方式
                     text_rotation=0,           #旋转文本
                     wrap_text=False,           #自动换行
                     shrink_to_fit=False,       #收缩到合适尺寸
                     indent=0)                  #缩进
number_format = 'General'                       #数字格式
protection = Protection(locked=True,            #保护？锁定
                         hidden=False)          #保护？隐藏 


from openpyxl import Workbook,load_workbook
filename="filtered.xlsx"
wb=load_workbook(filename)
ws=wb.active
ws["c5"].border=border#:B15
ws["E5"].border=border#:G8
left, right, top, bottom = [Side(style='double', color='000000')] * 4
ws['g8'].border = Border(left=left, right=right, top=top, bottom=bottom)
wb.save(filename)

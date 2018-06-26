'''用python实现excel(.xlsx)按工作表分别复制到另一份.xlsx中'''
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Border,Side

def d_copysheets(loadfl,loadsheets,savefl,savesheets):
    """loadsheets,savesheets为sheet名的列表"""
    assert len(loadsheets)==len(savesheets)
    try:
        wb1=load_workbook(filename=loadfl)
        wb2=load_workbook(filename=savefl)
        for i,sheet in enumerate(loadsheets):
            ws1=wb1[sheet]
            ws2=wb2[savesheets[i]]
            #删除原表值
            for i,row in enumerate(ws2.iter_rows()):
                for j,cell in enumerate(row):
                    ws2.cell(row=i+1, column=j+1, value='')
            #开始复制
            for i,row in enumerate(ws1.iter_rows()):
                for j,cell in enumerate(row):
                    ws2.cell(row=i+1, column=j+1, value=cell.value)
    except FileNotFoundError as err:
        return print('文件名错误:',err)
    except KeyError as err:
        return print('工作表Sheet名错误:',err)
    else:
        wb2.save(savefl)
        wb2.close()
        wb1.close()
        return print('已将',loadfl,loadsheets,'\n复制到',savefl,savesheets,'\n','=='*10)

def d_rang_border(loadfl,loadsheets,range_BOR,border_st=None):
    if not border_st:
        thin = Side(border_style="thin", color="000000")
        border_st=Border(top=thin, left=thin, right=thin, bottom=thin)
    try:
        wb1=load_workbook(filename=loadfl)
        rows=wb1[loadsheets][range_BOR]
        for row in rows:
            for cell in row:
                cell.border=border_st
    except FileNotFoundError as err:
        return print('文件名错误:',err)
    except KeyError as err:
        return print('工作表Sheet名错误:',err)
    except ValueError as err:
        return print('范围设置错误:',err)
    else:
        wb1.save(loadfl)
        wb1.close()
        return print('已将',loadfl,'\n',loadsheets,'\n中的',range_BOR,'单元格设置好\n','=='*10)

def main():
    d_cmd=input('1:自动复制到三个挂起文件；\n2:手动输入;\n输入其他无效,请输入：')
    if d_cmd=='1' :
        if input('确认输入"ok"回车后，开始自动复制：')=='ok':
            print('请稍后。。。')
            d_copysheets(\
                r'E:\2017工作-ET\开网优化\进度统计\统计工具\access导出文件夹\查询4-1 本周清单_除异常工单.xlsx',\
                ['查询4_1_本周清单_除异常工单'],\
                r'E:\2017工作-ET\开网优化\进度统计\统计工具\access导出文件夹\输出模板\2.LTE开网优化状态库挂起工单2018xxxx.xlsx',\
                ['开网优化挂起工单'])
            d_copysheets(\
                r'E:\2017工作-ET\开网优化\进度统计\统计工具\access导出文件夹\查询2-123 建设类_本周清单.xlsx',\
                ['查询2_1_建设类_本周清单', '查询2_2_建设类_新增', '查询2_3_建设类_解除'],\
                r'E:\2017工作-ET\开网优化\进度统计\统计工具\access导出文件夹\输出模板\开网优化项目_建设类_挂起工单清单2018xxxx.xlsx',\
                ['开网优化建设类挂起工单', '本周新增挂起工单', '本周已解除挂起工单'])
            d_copysheets(\
                r'E:\2017工作-ET\开网优化\进度统计\统计工具\access导出文件夹\查询3-123 维护类_本周清单.xlsx',\
                ['查询3_1_维护类_本周清单', '查询3_2_维护类_新增', '查询3_3_维护类_解除'],\
                r'E:\2017工作-ET\开网优化\进度统计\统计工具\access导出文件夹\输出模板\开网优化项目_维护类_挂起工单清单2018xxxx.xlsx',\
                ['开网优化维护类挂起工单', '本周新增挂起工单', '本周已解除挂起工单'])
            d_rang_border(r'E:\2017工作-ET\开网优化\进度统计\统计工具\access导出文件夹\输出模板\2.LTE开网优化状态库挂起工单2018xxxx.xlsx',\
                '统计','a1:h24')
    elif d_cmd=='2':
        loadfl=input('输入源文件（.xlsx）：')
        loadsheets=input('输入源工作表列表(如："Sheet1,Sheet2")：').split(',')
        savefl=input('输入保存文件（.xlsx）：')
        savesheets=input('输入保存的工作表列表(如："Sheet1,Sheet2")：').split(',')
        if input('确认输入"ok"回车后，开始操作：')=='ok':
            print('请稍后。。。')
            d_copysheets(loadfl,loadsheets,savefl,savesheets)
    input('操作完成，按任意键退出:')
    return

if __name__ == '__main__':
    main()
    

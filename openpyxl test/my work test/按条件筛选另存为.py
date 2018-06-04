'''用python实现excel(.xlsx)字段筛选，然后另存为一个新文档'''
from openpyxl import load_workbook
from openpyxl import Workbook

def filterdata(loadfl,con_dict,savefl='new_excel.xlsx'):
    '''con_dict:筛选表头条件的字典参数'''
    wb1=load_workbook(filename=loadfl)
    ws1=wb1.active
    wb2=Workbook(write_only=True)
    ws2=wb2.create_sheet()
    ws2.append(ws1[1])        #加入第一行表头
    site_dict={i:v.value for i,v in enumerate(ws1[1]) if v.value in con_dict}        #将表头条件字典转换为列数的条件字典
    for row1 in ws1.rows:
        if conditionf(row1,con_dict,site_dict):
            ws2.append(row1)
    wb2.save(savefl)
    wb2.close()
    wb1.close()
    return print('操作完成')

def conditionf(wsrow,con_dict,site_dict):
    for each in site_dict:
        if not wsrow[each].value in con_dict[site_dict[each]]:
            return False
    return True

def main():
    loadfile='a.xlsx'
    savefile='b.xlsx'
    con_dict={
        'a':['lte'],
        'b':['c']
    }
    filterdata(loadfile,con_dict,savefile)    
    return


if __name__ == '__main__':
    main()

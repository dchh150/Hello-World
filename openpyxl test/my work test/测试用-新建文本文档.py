from openpyxl import load_workbook
from openpyxl import Workbook
loadfl='a - 副本.xlsx'
wb1=load_workbook(filename=loadfl)
ws1=wb1.active
for i,row in enumerate(ws1.iter_rows()):
    for j,cell in enumerate(row):
        ws1.cell(row=i+1, column=j+1, value='')
wb1.save(loadfl)
wb1.close()
<<<<<<< HEAD
#测试0617-----06174-0000

=======
#测试0617-----06174
#测试0617-----06173-112
>>>>>>> parent of 4ce2b7b... 更新测试
